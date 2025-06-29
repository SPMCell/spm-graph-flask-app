from datetime import datetime
import datetime
import json
import difflib
import os
import pandas as pd
import xlsxwriter
from config import CHART_FILE
from config import EXCEL_FILE


def load_station_data(json_file='stations.json'):
    with open(json_file, 'r') as f:
        return json.load(f)


def process_train_data(df, base_km, segment_stations):
    """
    Processes train data by computing absolute km values and then applying a reset
    adjustment if a reset station is detected.

    Step 1: Uses the km value of the starting station and the next station to determine
    whether the segment is ascending or descending.
       - If ascending, compute: absolute_km = base_km + Distance_KM
       - If descending, compute: absolute_km = base_km - Distance_KM

    Step 2: Checks whether any station in segment_stations has a 'reset' flag (e.g., RRB).
            If so, it locates the corresponding 'reset' row based on:
              - absolute_km being within (approach_km + tolerance), and
              - Loco_Speed == 0.
            Then, it uses the reset station’s km (base_km_reset) and the Distance_KM value
            at that reset row (D_reset) to recalculate all subsequent absolute_km values
            using:
              new_absolute_km = base_km_reset - (current_Distance_KM - D_reset)

    Parameters:
      df              : DataFrame containing train data. (It should contain columns like
                        "Distance_KM", "Loco_Speed", etc.)
      base_km         : float, the km value of the starting station.
      segment_stations: dict, the segment-specific station data.

    Returns:
      Updated DataFrame with computed 'absolute_km' column.
    """
    # -------------------------------
    # Step 1. Initial absolute_km Calculation
    # -------------------------------
    station_keys = list(segment_stations.keys())
    km_values = [segment_stations[station]["km"] for station in station_keys]

    # Validate that there are at least two stations
    if len(km_values) < 2:
        raise ValueError("Segment does not have enough stations to determine direction.")

    # Determine whether the segment is ascending or descending
    if km_values[0] < km_values[1]:
        # Ascending order: add the traveled distance
        df["absolute_km"] = base_km + df["Distance_KM"]
    else:
        # Descending order: subtract the traveled distance
        df["absolute_km"] = base_km - df["Distance_KM"]

    # Return the DataFrame with the initial absolute_km and any reset corrections applied.
    return df


def extract_time_only(time_val):
    """
    Extracts only the time portion (HH:MM:SS) from a given date-time value.

    If time_val is a pandas Timestamp or datetime, formats it as HH:MM:SS.
    If it's already a time-only string (e.g., "03:25:48"), it returns it unchanged.
    Otherwise, if it's a full datetime string, it parses it and returns only the time.
    """
    # If already a pandas Timestamp or datetime, format directly.
    if isinstance(time_val, (pd.Timestamp, datetime)):
        return time_val.strftime('%H:%M:%S')

    # If it's a string, remove any leading/trailing whitespace.
    if isinstance(time_val, str):
        time_val = time_val.strip()
        # Check if the string is already only time (e.g., "03:25:48")
        # A simple check: length of 8 and contains two colons.
        if len(time_val) == 8 and time_val.count(':') == 2:
            return time_val

        # Otherwise, try to parse as full datetime string.
        try:
            dt = datetime.strptime(time_val, '%d-%m-%Y %H:%M:%S')
            return dt.strftime('%H:%M:%S')
        except Exception as e:
            print("Error in extract_time_only:", e)
            return time_val

    # If none of the above, simply return the value as string.
    return str(time_val)


def get_best_pilot_match(user_input, json_file="loco_pilots.json", cutoff=0.5):
    # Load the JSON file containing pilot data.
    with open(json_file, "r") as f:
        data = json.load(f)

    pilots = data.get("loco_pilots", [])
    # Create a list of pilot names (converted to uppercase for case-insensitive matching).
    pilot_names = [pilot["name"].upper() for pilot in pilots]

    # Convert user input to uppercase.
    user_input_upper = user_input.upper()

    # Get the best close match.
    matches = difflib.get_close_matches(user_input_upper, pilot_names, n=1, cutoff=cutoff)

    if matches:
        best_match = matches[0]
        for pilot in pilots:
            if pilot["name"].upper() == best_match:
                return pilot
    return None


def get_valid_date(date_str=None, prompt="Enter Date (DD-MM-YYYY): "):
    while True:
        if not date_str:
            date_str = input(prompt).strip()
        try:
            datetime.datetime.strptime(date_str, "%d-%m-%Y")
            return date_str
        except ValueError:
            print("Invalid date format. Expected DD-MM-YYYY.")
            date_str = None


def get_valid_time(time_str=None, prompt="Enter Time (HH:MM or HH:MM:SS): "):
    # If no time_str is provided, use interactive prompt.
    while True:
        # If time_str is None or empty, prompt the user
        if not time_str:
            time_str = input(prompt).strip()
        # If user entered HH:MM, append :00
        if time_str.count(":") == 1:
            time_str += ":00"
        try:
            datetime.datetime.strptime(time_str, "%H:%M:%S")
            return time_str
        except ValueError:
            # In interactive use, print error and prompt again.
            print("Invalid time format. Expected HH:MM or HH:MM:SS.")
            time_str = None  # Clear and try again in interactive mode.
            # In web mode, you might prefer raising an error instead,
            # but you can also choose to just return a default value or flash a message.
            # For now, we mimic interactive behavior.
            # You could also raise an exception if you want to handle it in your route.


def get_route_segment(route_stations, from_station, to_station):
    """
    Given a dictionary of route_stations (assumed to be in the desired order)
    and the desired start and end station codes, return a dictionary containing
    only the stations in that segment.
    """
    from_station = from_station.upper()
    to_station = to_station.upper()
    keys = list(route_stations.keys())
    try:
        start_index = keys.index(from_station)
        end_index = keys.index(to_station)
    except ValueError:
        print("Error: One of the specified stations is not found.")
        return {}

    # If order goes forward, return that slice.
    if start_index <= end_index:
        segment_keys = keys[start_index:end_index + 1]
    else:
        # If reversed, you could either decide that it’s an error
        # or return the slice in reverse order.
        segment_keys = keys[end_index:start_index + 1]
    return {k: route_stations[k] for k in segment_keys}


def get_stop_events(train_data, speed_threshold=0.1):
    """
    Extracts the first row of each contiguous block where speed is below or equal to the threshold.
    By default, we use speed_threshold=0.1 to treat any near-zero speeds as stops.

    Assumptions:
      - The DataFrame 'train_data' is sorted by time.
      - The speed column is "Loco_Speed".

    Returns:
      A DataFrame containing one row per stop event.
    """
    # Create a boolean mask where True indicates a stop (i.e. speed below or equal to threshold)
    is_stop = train_data["Loco_Speed"] <= speed_threshold
    # Shift the mask to see where a new stop event starts:
    # A new stop event is when is_stop is True and the previous row was not a stop.
    new_stop = is_stop & (~is_stop.shift(1, fill_value=False))
    stop_events = train_data[new_stop]
    return stop_events


def filter_close_stop_events(stop_events, merge_distance=1.5):
    """
    Filters the stop_events DataFrame so that if two (or more) stops happen within
    merge_distance km (i.e. the difference in their 'absolute_km' values), only the later stop
    is retained. This function assumes that stop_events is sorted by datetime.

    Parameters:
       stop_events : DataFrame containing stop event rows (with at least 'absolute_km' and 'datetime').
       merge_distance : float, the maximum km difference within which stops are merged.

    Returns:
       A DataFrame of filtered stop events.
    """
    # Create an empty list to hold filtered rows.
    filtered_rows = []

    for idx, row in stop_events.iterrows():
        if not filtered_rows:
            # The very first event always gets added.
            filtered_rows.append(row)
        else:
            # Get the last accepted event.
            last_event = filtered_rows[-1]
            # Compare the absolute km values of the current event and the last event.
            # If they are within the merge_distance, replace the last event with the current one (later stop).
            if abs(row["absolute_km"] - last_event["absolute_km"]) < merge_distance:
                filtered_rows[-1] = row
            else:
                filtered_rows.append(row)

    # Convert the list of rows back to a DataFrame.
    return pd.DataFrame(filtered_rows)


def get_station_name_for_stop(segment_stations, detected_km, tolerance=2.5):
    """
    Given a dictionary of station data and a detected absolute kilometer value,
    return the station name whose km value is within a specified tolerance of the
    detected km. If multiple stations qualify, the one with the smallest difference
    is returned. If none fall within the tolerance window, returns None.

    Parameters:
      segment_stations: dict
          Dictionary mapping station names to details. Each value should be a dict
          containing at least the key "km" with its corresponding kilometer value.
      detected_km: float
          The absolute kilometer value detected from the train data.
      tolerance: float (default 2)
          Acceptable difference in kilometers between the detected km and the station's km.

    Returns:
      The station name (str) that best matches the detected kilometer (if within tolerance),
      or None if no match is found.
    """
    candidate_station = None
    min_diff = float("inf")

    for station, details in segment_stations.items():
        try:
            station_km = float(details["km"])
        except (KeyError, ValueError):
            # Skip station if km value is missing or invalid.
            continue

        diff = abs(station_km - detected_km)
        if diff <= tolerance and diff < min_diff:
            min_diff = diff
            candidate_station = station

    return candidate_station


def adjust_absolute_km(df_train, segment_stations, tolerance=2.5, offset_threshold=0.3):
    """
    Adjust the 'absolute_km' column of df_train based on stop events.
    If the TLGP–ASK section is detected (i.e. 'RRB' exists in segment_stations),
    the DataFrame is split into two parts (pre-RRB and post-RRB) and processed separately.
    Otherwise, the usual adjustment is made.
    """
    # [Normal processing using your existing logic]
    stop_events = get_stop_events(df_train, speed_threshold=0.1)
    if stop_events.empty:
        print("No stop events detected; no km adjustment applied.")
        return df_train
    corrected_df = df_train.copy()
    current_correction = 0.0
    for idx, stop_event in stop_events.iterrows():
        event_time = stop_event["datetime"]
        matching_row = corrected_df[corrected_df["datetime"] == event_time]
        if matching_row.empty:
            continue
        detected_km = matching_row.iloc[0]["absolute_km"]
        station = get_station_name_for_stop(segment_stations, detected_km, tolerance=tolerance)
        if station is None:
            continue
        try:
            official_km = float(segment_stations[station]["km"])
        except (KeyError, ValueError):
            continue
        new_offset = detected_km - official_km
        if abs(new_offset) >= offset_threshold:
            print(f"Stop at {event_time}: station '{station}' gives offset {new_offset:+.3f} km.")
            correction_mask = corrected_df["datetime"] >= event_time
            corrected_df.loc[correction_mask, "absolute_km"] -= new_offset
            # Optionally, update the current_correction tracker (if needed for debugging or chaining)
            current_correction += new_offset
        else:
            print(f"Stop at {event_time}: offset {new_offset:+.3f} km within threshold.")
    df_train = corrected_df  # Use the corrected values from stop events
    station_keys = list(segment_stations.keys())
    for station in station_keys:
        if segment_stations[station].get("reset", False):  # Detect a reset station
            print(f"Reset station '{station}' detected. Changing absolute km calculation.")
            # Use a tolerance suitable for the reset logic (can be different from stop event tolerance)
            reset_tol = 2.0

            # Find the reset index: where 'absolute_km' is within (approach_km + reset_tol)
            # and where Loco_Speed is 0 (i.e. the train is stopped)
            reset_idx = df_train[
                (df_train["absolute_km"] <= (segment_stations[station]["approach_km"] + reset_tol)) &
                (df_train["Loco_Speed"] == 0)
                ].index.min()
            print("Reset index found:", reset_idx)
            if reset_idx is not None:
                base_km_reset = float(segment_stations[station]["km"])
                # Capture the Distance_KM value at the reset point.
                D_reset = df_train.loc[reset_idx, "Distance_KM"]
                print(f"At reset index {reset_idx}: base_km_reset = {base_km_reset}, D_reset = {D_reset}")

                # For all rows from reset_idx onward, adjust absolute_km:
                df_train.loc[reset_idx:, "absolute_km"] = base_km_reset - (
                            df_train.loc[reset_idx:, "Distance_KM"] - D_reset)
            break  # Apply reset logic only once
    return corrected_df


def load_and_prepare_data(excel_file=EXCEL_FILE, start_datetime=None, end_datetime=None):
    import openpyxl
    from io import BytesIO
    """
    Loads the Excel file and adapts to one of four formats using openpyxl in read-only mode.
    """

    # --------------------------------------------
    # Step 1: Read headers manually using openpyxl in read-only mode
    # --------------------------------------------
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else '' for h in next(rows)]
    except Exception as e:
        print(f"[ERROR] Failed to read headers from Excel file: {e}")
        return pd.DataFrame()

    separate_dt = any("date" in col.lower() or "/" in col for col in headers)
    file_type = "separate" if separate_dt else "combined"

    # --------------------------------------------
    # Step 2: Reconstruct the DataFrame manually from rows
    # --------------------------------------------
    data = [list(row) for row in rows]
    df = pd.DataFrame(data, columns=headers)
    df.columns = df.columns.str.strip()

    rename_map = {
        "Date dd/mm/yy": "Date dd/mm/yy",
        "DD/MM/YY": "Date dd/mm/yy",
        "Time hh:mn:ss": "Time hh:mn:ss",
        "hh:mm:ss": "Time hh:mn:ss",
        "Loco Speed (Kmph)": "Loco_Speed",
        "SPD(KMPH)": "Loco_Speed",
        "Distance (meters)": "Distance (Mtrs)",
        "Distance(Mtrs)": "Distance (Mtrs)",
        "Distance (Mtrs)": "Distance (Mtrs)",
        "DIST(Mtrs)": "Distance (Mtrs)",
        "BK_PIPE_PR(PSI)": "Brake Pipe Pr. (PSI)",
        "BP (Kg/cm2)": "Brake Pipe Pr. (Kg/Cm2)",
        "Brake Pipe Pr. (Kg/Cm2)": "Brake Pipe Pr. (Kg/Cm2)",
        "Time": "Time hh:mn:ss",
        "Speed Km/hr": "Loco_Speed",
        "Distance Km": "Distance",
        "BPP psi": "Brake Pipe Pr. (PSI)"
    }
    df.rename(columns=rename_map, inplace=True)
    if "BPP psi" not in df.columns:
        df["BPP psi"] = None

    # --------------------------------------------
    # Step 3: Create a unified datetime column
    # --------------------------------------------
    if file_type == "separate":
        combined = df["Date dd/mm/yy"].astype(str).str.strip() + " " + df["Time hh:mn:ss"].astype(str).str.strip()
        df["datetime"] = pd.to_datetime(combined, format="%d/%m/%y %H:%M:%S", errors='coerce')
    else:
        dt_series = pd.to_datetime(df["Time hh:mn:ss"].astype(str).str.strip(), format="%d-%m-%Y %H:%M:%S", errors='coerce')
        if dt_series.isna().all():
            dt_series = pd.to_datetime(df["Time hh:mn:ss"].astype(str).str.strip(), format="%d/%m/%y %H:%M:%S", errors='coerce')
        df["datetime"] = dt_series

    print("=== Debug Info in load_and_prepare_data ===")
    print("Total rows before filtering:", df.shape[0])
    print("Min datetime in dataset:", df["datetime"].min())
    print("Max datetime in dataset:", df["datetime"].max())
    print(df["datetime"].head(10).dt.strftime("%d-%m-%Y %H:%M:%S"))

    if start_datetime and end_datetime:
        try:
            start_dt = pd.to_datetime(start_datetime.strip(), format="%d-%m-%Y %H:%M:%S", errors='raise')
            end_dt = pd.to_datetime(end_datetime.strip(), format="%d-%m-%Y %H:%M:%S", errors='raise')
        except Exception as e:
            raise e

        extended_start_dt = start_dt - pd.Timedelta(minutes=5)
        extended_end_dt = end_dt + pd.Timedelta(minutes=5)

        df = df[(df["datetime"] >= extended_start_dt) & (df["datetime"] <= extended_end_dt)]

    if df.empty:
        return df

    speed_threshold = 0.1
    df = df.sort_values("datetime").reset_index(drop=True)
    departure_rows = df[df["Loco_Speed"] >= speed_threshold]
    if not departure_rows.empty:
        first_moving_index = departure_rows.index[0]
        actual_departure_time = df.iloc[first_moving_index - 1]["datetime"] if first_moving_index > 0 else departure_rows.iloc[0]["datetime"]
        df = df[df["datetime"] >= actual_departure_time]

    for i in range(len(df) - 2, -1, -1):
        if df.iloc[i]["Loco_Speed"] > speed_threshold >= df.iloc[i + 1]["Loco_Speed"]:
            df = df[df["datetime"] <= df.iloc[i + 1]["datetime"]]
            break

    distance_cols = [col for col in df.columns if "distance" in col.lower()]
    if not distance_cols:
        raise ValueError("No Distance column found in the file.")

    distance_col = distance_cols[0]
    if "mtrs" in distance_col.lower() or "meters" in distance_col.lower():
        df['Distance_KM'] = df[distance_col].cumsum() / 1000.0
    elif "km" in distance_col.lower() or distance_col.strip().lower() == "distance":
        df['Distance_KM'] = df[distance_col]
    else:
        raise ValueError("Distance unit not recognized in header.")

    df['Distance_KM'] = df['Distance_KM'] - df['Distance_KM'].iloc[0]

    bp_cols = [col for col in df.columns if "brake pipe pr" in col.lower()]
    if not bp_cols:
        raise ValueError("No Brake Pipe Pressure column found.")
    bp_col = bp_cols[0]
    if "kg" in bp_col.lower():
        df['BP_kg_cm2'] = df[bp_col]
    elif "psi" in bp_col.lower() or bp_col.strip().lower() == "bpp":
        df['BP_kg_cm2'] = df[bp_col] * 0.0703
    else:
        raise ValueError("Brake Pipe Pressure unit not recognized in header.")

    speed_candidates = [col for col in df.columns if "loco speed" in col.lower()]
    if not speed_candidates:
        speed_candidates = [col for col in df.columns if "speed" in col.lower()]
    if speed_candidates:
        df.rename(columns={speed_candidates[0]: "Loco_Speed"}, inplace=True)
    else:
        raise ValueError("No speed column found.")

    return df

# A helper function alias for clarity (choosing a shorter name):
def lookup_stop_name(segment_stations, km, tolerance=3):
    return get_station_name_for_stop(segment_stations, km, tolerance)

def generate_excel_chart(df, output_excel_path, train_no, section, lp_name, segment_stations):
    """
    Writes the main data to Excel and creates a line chart. In addition, a helper column
    ("Stop Marker") is built on a row‐by‐row basis so that if a particular row corresponds
    to the first instance of a stop event (as determined by get_stop_events and deduplication),
    a 0 is written in that row (otherwise =NA()). A custom data label using the station name is then
    attached to that point. This causes the label to appear exactly where the train stops (i.e. along the x‑axis).
    """
    # Create the workbook and worksheet.
    workbook = xlsxwriter.Workbook(output_excel_path)
    worksheet = workbook.add_worksheet("Data")

    # Write main data headers in columns A–C.
    main_headers = ["Absolute KM", "Loco Speed", "BP (Kg/Cm^2)"]
    for col_num, header in enumerate(main_headers):
        worksheet.write(0, col_num, header)

    # First, process the stop events to determine which rows are the first instance of a stop.
    # Use your existing helper functions.
    stop_events = get_stop_events(df, speed_threshold=0.1)
    filtered_stop_events = filter_close_stop_events(stop_events, merge_distance=1.5)
    # Deduplicate so that only the first row per stop (based on Absolute KM) is kept.
    unique_stops = filtered_stop_events.drop_duplicates(subset=["absolute_km"], keep="first")
    # Create a set containing the DataFrame indices (not the worksheet row numbers) of stops.
    unique_stop_indices = set(unique_stops.index)

    # Prepare lists to hold custom labels for every row (for the stops series).
    # For rows that are non-stops, an empty dictionary (or {}) will be used.
    custom_labels = []

    # Write main data rows (starting at row 1 in the worksheet) and build a helper column.
    # We'll write the helper column ("Stop Marker") in column D.
    for row_num, (idx, row) in enumerate(df.iterrows(), start=1):
        worksheet.write(row_num, 0, row["absolute_km"])
        worksheet.write(row_num, 1, row["Loco_Speed"])
        bp_val = row.get("BP_kg_cm^2", row.get("BP_kg_cm2", None))
        worksheet.write(row_num, 2, bp_val)

        # For the Stop Marker column, if this row is the first instance of a stop, write 0 and record its label;
        # otherwise, write a formula that returns #N/A so nothing is plotted.
        if idx in unique_stop_indices:
            # Get the station name corresponding to this stop.
            station_name = lookup_stop_name(segment_stations, row["absolute_km"], tolerance=3)
            worksheet.write(row_num, 3, 0)
            custom_labels.append({'value': station_name, 'position': 'above'})
        else:
            # Write a formula that returns NA.
            worksheet.write_formula(row_num, 3, '=NA()')
            custom_labels.append({})  # no label for non-stop rows

    # Create the line chart.
    chart = workbook.add_chart({'type': 'line'})

    # Main series: Loco Speed.
    chart.add_series({
        'name': 'Loco Speed',
        'categories': ['Data', 1, 0, df.shape[0], 0],  # Categories from column A.
        'values': ['Data', 1, 1, df.shape[0], 1],  # Values from column B.
        'line': {'color': 'blue'},
    })

    # Second series: Brake Pipe Pressure (secondary y-axis).
    chart.add_series({
        'name': 'BP',
        'categories': ['Data', 1, 0, df.shape[0], 0],
        'values': ['Data', 1, 2, df.shape[0], 2],
        'y2_axis': True,
        'line': {'color': 'red'},
    })
    chart.set_y2_axis({
        'name': 'BP',
        'max': 5.5  # Set your desired maximum here.
    })

    # Stops series: Use the helper column (column D) for y-values.
    # Because the helper column covers every row, the x-axis alignment (categories) is exactly the same as the main series.
    # Custom data labels will show the station name only on those rows where a stop occurred.
    chart.add_series({
        'name': 'Stop Stations',
        'categories': ['Data', 1, 0, df.shape[0], 0],  # Use same categories as the main series.
        'values': ['Data', 1, 3, df.shape[0], 3],  # Stop Marker values from column D.
        'data_labels': {
            'custom': custom_labels,
            'font': {
                'name': 'Calibri',  # Change to your preferred font name.
                'size': 11,  # Adjust the font size as needed.
                'color': 'orange'  # Change the color if desired.
            }
        },
        'marker': {'type': 'square', 'size': 8, 'fill': {'color': 'green'}},
        'line': {'none': True},
    })

    # Set the chart title, axis labels, and other formatting.
    chart.set_title({
        'name': f"Tr.No. {train_no} {section}         {lp_name}",
        'font': {'size': 16},
        'layout': {'x': 0.2, 'y': 0.0}
    })
    chart.set_x_axis({'name': 'Distance in KM'})
    chart.set_y_axis({'name': 'Speed (kmph)'})
    chart.set_y2_axis({'name': 'BP (Kg/Cm^2)'})
    chart.set_size({'width': 900, 'height': 600})
    chart.set_legend({
        'position': 'top',
        'layout': {'x': 0.9, 'y': 0.0, 'width': 0.18, 'height': 0.05}
    })
    chart.set_plotarea({
        'layout': {'x': 0.05, 'y': 0.05, 'width': 0.90, 'height': 0.80}
    })
    chart.set_style(2)

    # Insert the chart and close workbook.
    worksheet.insert_chart('H2', chart)
    workbook.close()
    print(f"Excel chart generated as {CHART_FILE}")


def assign_stop_labels(worksheet, df, segment_stations):
    """
    Processes stop events using get_stop_events, deduplicates them
    (using the absolute km value) so that only the first instance for each stop is kept,
    and writes helper columns in the worksheet for the stops series.

    The helper columns (in columns D, E, F) are:
      D: Stop Abs KM
      E: Stop Speed (forced to 0 so that stops are drawn on the x-axis)
      F: Stop Station (the station name)

    Returns:
      The list of station names (stop labels) for the deduplicated stop events.
    """
    # 1. Get stop events using your existing function.
    stop_events = get_stop_events(df, speed_threshold=0.1)

    # 2. Deduplicate stops based on the 'absolute_km' value.
    unique_stops = stop_events.drop_duplicates(subset=["absolute_km"], keep="first")

    # 4. Build lists: use the absolute km values and ignore the speed (force it to 0).
    stop_abs_km = unique_stops["absolute_km"].tolist()
    stop_speed = [0] * len(stop_abs_km)  # Force the speed to 0 for all stops.
    stop_labels = [
        lookup_stop_name(segment_stations, km, tolerance=3) or "Unknown"
        for km in stop_abs_km
    ]

    # 5. Write helper columns into the worksheet.
    if stop_labels:
        worksheet.write(0, 3, "Stop Abs KM")  # Column D
        worksheet.write(0, 4, "Stop Speed")  # Column E
        worksheet.write(0, 5, "Stop Station")  # Column F
        for i, (km, label) in enumerate(zip(stop_abs_km, stop_labels), start=1):
            worksheet.write(i, 3, km)  # Write the x-axis value.
            worksheet.write(i, 4, 0)  # Write the forced 0 for stop speed.
            worksheet.write(i, 5, label)  # Write the station label.

    return stop_labels


def convert_to_xlsx(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        try:
            # Use xlrd to read .xls files
            df = pd.read_excel(file_path, engine="xlrd")
        except Exception as e:
            raise Exception(f"Error converting file: {e}")
        # Create a new filename with .xlsx extension
        new_file_path = os.path.splitext(file_path)[0] + ".xlsx"
        # Save the DataFrame as a .xlsx file using openpyxl
        df.to_excel(new_file_path, index=False, engine="openpyxl")
        return new_file_path
    return file_path